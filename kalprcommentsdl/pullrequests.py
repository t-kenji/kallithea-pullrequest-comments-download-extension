# -*- coding: utf-8 -*-

import os
import gettext
import logging

from tempfile import mkstemp
from mako.template import Template
from pylons import response, tmpl_context as c
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.writer.excel import save_virtual_workbook

import kallithea
from kallithea.lib import helpers as h
from kallithea.lib.extensions import (
    IRoute, ITemplatePullrequests,
)
from kallithea.lib.vcs.utils import safe_str
from kallithea.model.db import PullRequest
from kallithea.model.comment import ChangesetCommentsModel

log = logging.getLogger(__name__)

localedir = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'i18n')
translate = gettext.translation('kalprcommentsdl', localedir, ['ja'], fallback=True)
_ = translate.ugettext

class CommentsDownload(IRoute, ITemplatePullrequests):

    log.info('CommentsDownload: class')

    def __init__(self):
        pass

    def make_map(self, config, rmap):

        from kallithea.lib.utils import (is_valid_repo, is_valid_repo_group,
                                         get_repo_by_id)

        def check_repo(environ, match_dict):
            """
            check for valid repository for proper 404 handling

            :param environ:
            :param match_dict:
            """
            repo_name = match_dict.get('repo_name')

            if match_dict.get('f_path'):
                #fix for multiple initial slashes that causes errors
                match_dict['f_path'] = match_dict['f_path'].lstrip('/')

            by_id_match = get_repo_by_id(repo_name)
            if by_id_match:
                repo_name = by_id_match
                match_dict['repo_name'] = repo_name

            return is_valid_repo(repo_name, config['base_path'])


        rmap.connect('pullrequest_export',
                     '/{repo_name:.*?}/pull-request-comment/{pull_request_id}/export/{fname}',
                     controller='pullrequests', action='export',
                     conditions=dict(function=check_repo, method=["GET"]))

    def add_property(self, ctx):
        data = dict(
            label=_('Download comments'),
            button=_('Downlaod as xlsx'),
        )
        url = '/kalprcommentsdl.pullrequests.commentdownload.html'
        template = u"""
<div class="label-summary">
  <label>{label}:</label>
</div>
<div class="input">
  <div>
    <a class="btn btn-small" href="${{h.url('pullrequest_export',repo_name=c.repo_name,pull_request_id=c.pull_request.pull_request_id,fname='pullrequest.xlsx')}}"><i class="icon-doc-inv"></i> {button}</a>
  </div>
</div>
""".format(**data)
        comment_download = Template(template)

        tmpl_lookup = kallithea.CONFIG['pylons.app_globals'].mako_lookup
        tmpl_lookup.put_string(url, template)
        return url

    def export(self, repo_name, pull_request_id, fname, **kwargs):
        ext = fname.split('.')[1]
        export_name = '{repo}-{pr_id}.{ext}'.format(repo=safe_str(repo_name.replace('/', '_')),
                                                    pr_id=safe_str(pull_request_id),
                                                    ext=safe_str(ext))
        fd, export_path = mkstemp()
        log.debug('Creating new temp export in {path}'.format(path=export_path))

        try:
            pr = PullRequest.get(pull_request_id)
            if repo_name != pr.other_repo.repo_name:
                raise RepositoryError
        except Exception as e:
            log.error(e)
            return _('Pull request #{id} not found').format(id=pull_request_id)
        cc_model = ChangesetCommentsModel()
        inline_comments = cc_model.get_inline_comments(
                            pr.org_repo_id,
                            pull_request=pull_request_id)
        file_comments = {}
        for f_path, lines in inline_comments:
            file_comments[f_path] = lines
        sorted_file_comments_by_name = sorted(file_comments.items(), key=lambda x:x[0], reverse=False)
        general_comments = cc_model.get_comments(pr.org_repo_id,
                                         pull_request=pull_request_id)

        wb = Workbook()
        ws = wb.create_sheet(_('comments'), 0)
        ws['A1'].value = _('File path')
        ws.column_dimensions['A'].width = 3.0
        ws['B1'].value = _('Comment ID')
        ws['C1'].value = _('Line no (old)')
        ws['D1'].value = _('Line no (new)')
        ws['E1'].value = _('Author')
        ws['F1'].value = _('Status')
        ws['G1'].value = _('Comment')
        ws.column_dimensions['G'].width = 60.0
        ws['H1'].value = _('Opinion')
        ws.column_dimensions['H'].width = 60.0
        ws['I1'].value = _('Retouch')
        ws['J1'].value = _('Priority')
        ws['K1'].value = _('Deadline')

        align_rot_90 = Alignment(text_rotation=90)
        align_wrap = Alignment(wrap_text=True)

        rows = 2
        for f_path, lines in sorted_file_comments_by_name:
            sorted_inline_comments_by_lineno = sorted(lines.iteritems(), key=lambda (line_no,comments):int(line_no[1:]), reverse=False)
            base_rows = rows
            for line_no, comments in sorted_inline_comments_by_lineno:
                for co in comments:
                    link = pr.url(canonical=True, anchor='comment-{id}'.format(id=co.comment_id))
                    ws['B{row}'.format(row=rows)].value = co.comment_id
                    ws['B{row}'.format(row=rows)].hyperlink = link
                    if co.line_no.startswith('o'):
                        ws['C{row}'.format(row=rows)].value = co.line_no[1:]
                    else:
                        ws['D{row}'.format(row=rows)].value = co.line_no[1:]
                    ws['E{row}'.format(row=rows)].value = co.author.username
                    if co.status_change:
                        ws['F{row}'.format(row=rows)].value = str(h.changeset_status_lbl(co.status_change[0].status))
                    ws['G{row}'.format(row=rows)].value = co.text.replace('@', '(at)')
                    ws['G{row}'.format(row=rows)].alignment = align_wrap
                    ws['H{row}'.format(row=rows)].alignment = align_wrap
                    rows += 1
            ws.merge_cells('A{start}:A{end}'.format(start=base_rows, end=rows-1))
            for i in range(rows-base_rows):
                ws['A{row}'.format(row=base_rows+i)].value = f_path
            ws['A{start}'.format(start=base_rows)].alignment = align_rot_90

        ws['A{row}'.format(row=rows)].value = 'General'
        base_rows = rows
        for co in general_comments:
            link = pr.url(canonical=True, anchor='comment-{id}'.format(id=co.comment_id))
            ws['B{row}'.format(row=rows)].value = co.comment_id
            ws['B{row}'.format(row=rows)].hyperlink = link
            ws['E{row}'.format(row=rows)].value = co.author.username
            if co.status_change:
                ws['F{row}'.format(row=rows)].value = str(h.changeset_status_lbl(co.status_change[0].status))
            ws['G{row}'.format(row=rows)].value = co.text.replace('@', '(at)')
            ws['G{row}'.format(row=rows)].alignment = align_wrap
            ws['H{row}'.format(row=rows)].alignment = align_wrap
            rows += 1
        ws.merge_cells('A{start}:A{end}'.format(start=base_rows, end=rows-1))
        for i in range(rows-base_rows):
            ws['A{row}'.format(row=base_rows+i)].value = 'General'
        ws['A{start}'.format(start=base_rows)].alignment = align_rot_90

        with os.fdopen(fd, 'wb') as s:
            s.write(save_virtual_workbook(wb))

        def get_chunked_export(export_path):
            stream = open(export_path, 'rb')
            while True:
                data = stream.read(16 * 1024)
                if not data:
                    break
                yield data
            stream.close()
            log.debug('Destroying temp export %s', export_path)
            os.remove(export_path)

        response.content_disposition = str('attachment; filename=%s' % (export_name))
        response.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return get_chunked_export(export_path)
